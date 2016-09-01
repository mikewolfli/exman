#!/usr/bin/python3
#-*- coding:utf-8 -*-
'''
Created on 2016年7月19日

@author: mikewolfli
'''
import os,sys
from tkinter import *
from tkinter import simpledialog
from tkinter import font
from tkinter import scrolledtext
from tkinter import messagebox
from tkinter import filedialog
import tkinter.ttk as ttk
#import pandas as pd
#from pandastable import Table, TableModel
from ldap3 import Server, Connection, SIMPLE, SYNC, ALL, SASL, NTLM
from dataset import * 
import logging 
import datetime
from openpyxl import *
from openpyxl import writer
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import xlrd
import xlwt
from xlutils.copy import copy
#from treelib import Tree, Node
from decimal import Decimal
import pyrfc
import threading
import base64
from configparser import ConfigParser

NAME = 'EDS非标物料处理'
PUBLISH_KEY=' A ' #R - release , B - Beta , A- Alpha
VERSION = '0.1.0'

logger = logging.getLogger()

def cur_dir():
    #获取脚本路径
    path = sys.path[0]
    #判断为脚本文件还是py2exe编译后的文件，如果是脚本文件，则返回的是脚本的目录，
    #如果是py2exe编译后的文件，则返回的是编译后的文件路径
    if os.path.isdir(path):
        return path
    elif os.path.isfile(path):
        return os.path.dirname(path)

def value2key(dic, value):
    if not isinstance(dic, dict):
        return None
    
    for key, val in dic.items():
        if val == value:
            return key
    
    return None

def date2str(dt_s):
    if not isinstance(dt_s, datetime.datetime):
        return None
    else:
        return dt_s.strftime("%Y-%m-%d") 

def datetime2str(dt_s):
    if not isinstance(dt_s, datetime.datetime):
        return None
    else:
        return dt_s.strftime("%Y-%m-%d %H:%M:%S") 
    
def str2date(dt_s):
    if dt_s is None or (len(dt_s)==0 and isinstance(dt_s, str)):
        return None
    else:
        return datetime.datetime.strptime(dt_s , '%Y-%m-%d')
    
def str2datetime(dt_s):
    if dt_s is None or (len(dt_s)==0 and isinstance(dt_s, str)) :
        return None
    else:
        return datetime.datetime.strptime(dt_s, "%Y-%m-%d %H:%M:%S") 

def none2str(val):
    if not val:
        return ''
    else:
        return str(val).strip()
    
def cell2str(val):
    if (val is None) or (val=='N/A') or (val=='N') or (val=='无'):
        return ''
    else:
        return str(val).strip()

def tree_level(val):
    l = len(val)
    if l==0:
        return 0
    
    r=1
    for i in range(l):
        if int(val[i])>0:
            return r
        elif int(val[i])==0:
            r+=1
            
    return r

def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))

class TextHandler(logging.Handler):
    """This class allows you to log to a Tkinter Text or ScrolledText widget"""
    def __init__(self, text):
        # run the regular Handler __init__
        logging.Handler.__init__(self)
        # Store a reference to the Text it will log to
        self.text = text

    def emit(self, record):
        self.formatter = logging.Formatter('%(asctime)s-%(levelname)s : %(message)s')
        msg = self.format(record)
        def append():
            self.text.configure(state='normal')          
            self.text.insert(END, msg+"\n")
            self.text.configure(state='disabled')
            # Autoscroll to the bottom
            self.text.yview(END)
        # This is necessary because we can't modify the Text from other threads
        self.text.after(0, append)# Scroll to the bottom
        
login_info ={'uid':'','pwd':'','status':False,'perm':'000000','plant':'2101'}
        
class LoginForm(Toplevel):
    def __init__(self, parent, title=None):
        Toplevel.__init__(self, parent)

        self.withdraw()
        if parent.winfo_viewable():
            self.transient(parent)

        if title:
            self.title(title)

        self.parent = parent
        body = Frame(self)
        self.initial_focus = self.body(body)
        body.pack(padx=5, pady=5)

        self.buttonbox()

        if not self.initial_focus:
            self.initial_focus = self

        self.protocol("WM_DELETE_WINDOW", self.cancel)

        center(self)
        #if self.parent is not None:         
            #self.geometry("+%d+%d" % (parent.winfo_rootx()+50,
            #                         parent.winfo_rooty()+50))

        self.deiconify() # become visible now

        self.initial_focus.focus_set()

        # wait for window to appear on screen before calling grab_set
        self.wait_visibility()
        
        self.grab_set()
        
        self.wait_window(self) 
            
    def body(self, master):
        self.label1 = Label(master,text='用户名')
        self.label1.grid(row=0, column=0, sticky=W)
        self.uid_entry= Entry(master)
        self.uid_entry.grid(row=0, column=1, columnspan=2,  sticky=EW)       
        self.label2 = Label(master, text='密码')
        self.label2.grid(row=1, column=0, sticky=W)
        self.pwd_entry = Entry(master, show='*')
        self.pwd_entry.grid(row=1, column=1, columnspan=2, sticky=EW)
        self.msg_str=StringVar()
        self.label_message = Label(master, textvariable=self.msg_str).grid(row=2, column=0, columnspan=3, sticky=W)
        return self.uid_entry

    def buttonbox(self):
        box = Frame(self)

        w = Button(box, text="登陆", width=10, command=self.ok, default=ACTIVE)
        w.pack( side=LEFT, padx=5, pady=5)
        w = Button(box, text="取消", width=10, command=self.cancel)
        w.pack(side=LEFT, padx=5, pady=5)

        self.bind("<Return>", self.ok)
        self.bind("<Escape>", self.cancel)

        box.pack()
        
    def validate(self):  
        login_info['uid'] = self.uid_entry.get()
        login_info['pwd'] = self.pwd_entry.get()
        
        if len(login_info['uid'].strip())==0:
            self.msg_str.set('用户名不能为空')
            self.initial_focus=self.uid_entry
            self.initial_focus.focus_set()
            return 0
            
        s = Server('tkeasia.com', get_info=ALL)
        c = Connection(s, user="tkeasia\\"+login_info['uid'], password=login_info['pwd'], authentication=NTLM)
        ret = c.bind()
        if ret:
            login_info['status'] = True
            self.get_permission()
            self.log_login()
            
            logger.info(login_info['uid']+"登陆系统...")
            return 1
        else:
            self.msg_str.set('登陆失败！')
            messagebox.showerror('错误', c.last_error)
            return 0
        
    def destroy(self):
        self.initial_focus = None
        Toplevel.destroy(self)    
        
    def cancel(self, event=None):
        if self.parent is not None:
            self.parent.focus_set()
        self.destroy()
        
        if pg_db.get_conn():
            pg_db.close()
            
        if mbom_db.get_conn():          
            mbom_db.close()
            
        sys.exit()
        
    def ok(self, event=None):
        if not self.validate():
            self.initial_focus.focus_set() # put focus back
            return
        self.withdraw()
        self.update_idletasks()        
        self.destroy()
        
    def log_login(self):
        if not mbom_db.get_conn():
            mbom_db.connect()
              
        login_record = login_log.select().where((login_log.employee==login_info['uid'])&(login_log.log_status==True))
        if len(login_record) >0:   
            log_loger = login_log.update(log_status=False).where((login_log.employee==login_info['uid'])&(login_log.log_status==True))
            log_loger.execute()
        
        
        log_loger = login_log.insert(employee=login_info['uid'], log_status=True, login_time=datetime.datetime.now())
        log_loger.execute()
    
    def get_permission(self):
        try:
            perm = op_permission.get(op_permission.employee==login_info['uid'])
            login_info['perm']= perm.perm_code
        except op_permission.DoesNotExist:
            pass

mat_heads = ['位置号','物料号','中文名称','英文名称','图号','数量','单位','材料','重量','备注']
mat_keys = ['st_no','mat_no','mat_name_cn','mat_name_en','drawing_no','qty','mat_unit','mat_material','part_weight','comments']

mat_cols = ['col1','col2','col3','col4','col5','col6','col7','col8','col9','col10']

def dict2list(dict):
    li = []
    
    for i in range(len(mat_heads)):
        li.append(dict[mat_heads[i]])
    
    return li

#threads=[]
threadLock = threading.Lock()
class refresh_thread(threading.Thread):
    def __init__(self, frame, typ=None):
        threading.Thread.__init__(self)
        self.frame=frame
        self.type=typ

    def run(self):
        threadLock.acquire()
        self.frame.check_in_sap()
        threadLock.release()  

class mainframe(Frame):
    '''
    mat_list = {1:{'位置号':value,'物料号':value, ....,'标判断':value},.....,item:{......}}
    bom_tree : 物料BOM的树形结构, 以key为节点,保存om树型结构如下:
    0
    ├── 1
    │   └── 3
    └── 2
    '''
    hibe_mats=[]
    no_need_mats=[]
    mat_list = {} #从文件读取的文件列表，以 数字1，2，...为keys
    bom_items = [] #存储有下层BOM的节点，treeview 控件的节点
    mat_items = {} #以物料号为key,存储涉及BOM的物料清单 ，包括下层物料。
    #treeview本身是树形结构，无需在重新构建树形model
    #bom_tree = Tree()
    #bom_tree.create_node(0,0)
    mat_pos = 0 #配合mat_list的的数量
    mat_tops={} #发运层物料字典，key为物料号，value是struct code 和revision列表
    nstd_mat_list=[] #非标物料列表
    sap_thread = None
    def __init__(self,master=None):
        Frame.__init__(self, master)
        self.pack()
        
        self.createWidgets()
        
    def createWidgets(self):
        '''
        self.find_mode = StringVar()
        self.find_combo = ttk.Combobox(self,textvariable = self.find_mode)
        self.find_combo['values'] = ('列出物料BOM结构','查找物料的上层','查找物料关联项目','查找项目关联物料')
        self.find_combo.current(0)
        self.find_combo.grid(row =0,column=0, columnspan=2,sticky=EW)
        '''
        self.find_label = Label(self, text='请输入头物料号查找',anchor='w') 
        self.find_label.grid(row=0, column=0, columnspan=2, sticky=EW)
              
        self.find_var = StringVar()
        self.find_text = Entry(self, textvariable=self.find_var)
        self.find_text.grid(row=1, column=0, columnspan=2, sticky=EW)
        self.find_text.bind("<Return>", self.search)
        
        self.version_label = Label(self,text='物料版本', anchor ='w')
        self.version_label.grid(row=0, column=2, columnspan=2, sticky=EW)
        
        self.version_var = StringVar()
        self.version_text = Entry(self, textvariable=self.version_var)
        self.version_text.grid(row=1, column=2, columnspan=2, sticky=EW)
        self.version_text.bind("<Return>", self.search)
                
        self.st_body = Frame(self)
        self.st_body.grid(row=0, column=4,sticky=NSEW)
        
        self.import_button = Button(self.st_body, text='文件读取')
        self.import_button.pack(side='left')
        self.import_button['command'] = self.excel_import
        '''
        self.check_sap = Button(st_body, text='非标物料系统比对')
        self.check_sap.pack(side='left')
        self.check_sap['command']= self.run_check_in_sap
        '''
        
        self.generate_nstd_list = Button(self.st_body, text='生成非标物料申请表')
        self.generate_nstd_list.pack(side='left')
        self.generate_nstd_list['command']=self.generate_app
               
        self.ie_body = Frame(self)
        self.ie_body.grid(row=1, column=4, columnspan=10, sticky=NSEW)
        
        self.import_bom_List = Button(self.ie_body, text='生成BOM导入表')
        self.import_bom_List.pack(side='left')
        self.import_bom_List['command']=self.import_bom_list
        
        self.ntbook = ttk.Notebook(self)        
        self.ntbook.rowconfigure(0, weight=1)
        self.ntbook.columnconfigure(0, weight=1)
        '''
                清单式显示不够直观，同时pandastable表操作速度太慢，故只使用树形结构
        list_pane = Frame(self)
        model = TableModel(rows=0, columns=0)
        for col in mat_heads:
            model.addColumn(col)
        model.addRow(1)
        
        self.mat_table = Table(list_pane, model, editable=False)
        self.mat_table.show()
        '''                
        tree_pane = Frame(self)
        self.mat_tree = ttk.Treeview(tree_pane, columns=mat_cols,selectmode='extended')
        style = ttk.Style()
        style.configure("Treeview", font=('TkDefaultFont', 10))
        style.configure("Treeview.Heading", font=('TkDefaultFont', 9))  
        self.mat_tree.heading('#0', text='')
        for col in mat_cols:
            i = mat_cols.index(col)
            if i==0:
                self.mat_tree.heading(col,text="版本号/位置号")
            else:
                self.mat_tree.heading(col,text=mat_heads[i])
        
        #('位置号','物料号','中文名称','英文名称','图号','数量','单位','材料','重量','备注')
        self.mat_tree.column('#0', width=80)
        self.mat_tree.column('col1', width=80, anchor='w')
        self.mat_tree.column('col2', width=100, anchor='w')
        self.mat_tree.column('col3', width=150, anchor='w')
        self.mat_tree.column('col4', width=150, anchor='w')
        self.mat_tree.column('col5', width=100, anchor='w')
        self.mat_tree.column('col6', width=100, anchor='w')
        self.mat_tree.column('col7', width=100, anchor='w')
        self.mat_tree.column('col8', width=150, anchor='w')
        self.mat_tree.column('col9', width=100, anchor='w')
        self.mat_tree.column('col10', width=300, anchor='w')  
               
        ysb = ttk.Scrollbar(tree_pane, orient='vertical', command=self.mat_tree.yview)
        xsb = ttk.Scrollbar(tree_pane, orient='horizontal', command=self.mat_tree.xview)        
        ysb.grid(row=0, column=2, rowspan=2, sticky='ns')
        xsb.grid(row=2, column=0, columnspan=2, sticky='ew')  
        
        self.mat_tree.configure(yscroll=ysb.set, xscroll=xsb.set)
        self.mat_tree.grid(row=0, column=0, rowspan=2, columnspan=2, sticky='nsew')
        tree_pane.rowconfigure(1, weight=1)
        tree_pane.columnconfigure(1, weight =1)
        
        #self.ntbook.add(list_pane, text='BOM清单', sticky=NSEW)
        self.ntbook.add(tree_pane, text='BOM树形结构', sticky=NSEW) 
        
        log_pane = Frame(self)
        
        self.log_label=Label(log_pane)
        self.log_label["text"]="操作记录"
        self.log_label.grid(row=0,column=0, sticky=W)
        
        self.log_text=scrolledtext.ScrolledText(log_pane, state='disabled')
        self.log_text.config(font=('TkFixedFont', 10, 'normal'))
        self.log_text.grid(row=1, column=0, columnspan=2, sticky=EW)
        log_pane.rowconfigure(1,weight=1)
        log_pane.columnconfigure(1, weight=1)
        
        self.ntbook.grid(row=2, column=0, rowspan=6, columnspan=6, sticky=NSEW)
        log_pane.grid(row=8, column=0,columnspan=6, sticky=NSEW)
              
        # Create textLogger
        text_handler = TextHandler(self.log_text)        
        # Add the handler to logger
        
        logger.addHandler(text_handler)
        logger.setLevel(logging.INFO) 
        
        self.rowconfigure(8, weight=1)
        self.columnconfigure(5, weight=1) 
    
    def display_widgets(self):
        if login_info['perm'][3]!='1' and login_info['perm'][3]!='9':
            self.st_body.grid_forget()
            
        if login_info['perm'][4]!='1' and login_info['perm'][4]!='9':
            self.ie_body.grid_forget()
                       
    def import_bom_list(self):
        if len(self.bom_items)==0:
            logger.warning('没有bom结构，请先搜索物料BOM')
            return
        
        if self.sap_thread.is_alive():
            messagebox.showinfo('提示','正在后台检查SAP非标物料，请等待完成后再点击!')
            return  
        
        if len(self.nstd_mat_list) != 0:
            logger.warning('此物料BOM中包含未维护进系统的物料，请等待其维护完成')
            return

        file_str=filedialog.asksaveasfilename(title="导出文件", initialfile="temp",filetypes=[('excel file','.xls')])
        if not file_str:
            return

        if not file_str.endswith(".xls"):
            file_str+=".xls"        

        temp_file = os.path.join(cur_dir(),'bom.xls')
        rb = xlrd.open_workbook(temp_file, formatting_info=True)
                
        wb= copy(rb)
        ws = wb.get_sheet(0)
        
        logger.info('正在生成文件'+file_str)
        i=4
        for it in self.bom_items:
            p_mat = self.mat_tree.item(it, 'values')[1]
            logger.info('正在构建物料'+p_mat+'的BOM导入清单...')
            p_name = self.mat_tree.item(it, 'values')[2]
            children = self.mat_tree.get_children(it)
            for child in children:
                value = self.mat_tree.item(child, 'values')
                c_mat = value[1]
                c_name = value[2]
                ws.write(i, 0, p_mat)
                ws.write(i, 1, p_name)
                ws.write(i, 5, c_mat)
                ws.write(i, 6, c_name)
                ws.write(i, 2, 2102)
                ws.write(i, 3, 1)
                if c_mat in self.hibe_mats:
                    ws.write(i, 4, 'N')
                else:
                    ws.write(i, 4, 'L')
                    ws.write(i, 14, 'X')
                    
                ws.write(i, 7, float(value[5]))
                i+=1
                        
        wb.save(file_str)
        logger.info(file_str+'保存完成!')
    
    def generate_app(self):
        if len(self.nstd_mat_list)==0:
            logger.warning('没有非标物料，无法生成非标物料申请表')
            return
        
        nstd_id = simpledialog.askstring('非标申请编号','请输入完整非标申请编号(不区分大小写)，系统将自动关联项目:')
        nstd_id = nstd_id.upper().strip()
        
        basic_info = self.get_rel_nstd_info(nstd_id)
        if not basic_info:
            logger.warning('非标申请：'+nstd_id+'在流程软件中未创建，请先创建后再生成非标物料申请表!')
            return
        
        file_str=filedialog.asksaveasfilename(title="导出文件", initialfile="temp",filetypes=[('excel file','.xlsx')])
        if not file_str:
            return

        if not file_str.endswith(".xlsx"):
            file_str+=".xlsx"
        
        if not self.create_nstd_mat_table(nstd_id, basic_info):
            logger.warning('由于非标物料均已经在其他非标申请中提交，故中止创建非标申请清单文件。')
            return
            
        temp_file = os.path.join(cur_dir(),'template.xlsx')
        logger.info('正在根据模板文件:'+temp_file+'生成申请表...')
        wb = load_workbook(temp_file) 
        temp_ws = wb.get_sheet_by_name('template')
        m_qty = len(self.nstd_mat_list)
        
        if m_qty%28 ==0:
            s_qty = m_qty/28
        else:
            s_qty = int(m_qty/28)+1
            
        for i in range(1, s_qty+1):
            ws = wb.copy_worksheet(temp_ws)     
            ws.sheet_state ='visible'
            ws.title = 'page'+str(i)
            self.style_worksheet(ws)
            
            ws.cell(row=5,column=1).value = 'Page '+str(i)+'/'+str(s_qty)
            logger.info('正在向第'+str(i)+'页填入物料数据...')
            self.fill_nstd_app_table(ws, i, nstd_id, basic_info,m_qty)
        
        wb.remove_sheet(temp_ws)
        if writer.excel.save_workbook(workbook=wb, filename=file_str):
            logger.info('生成非标物料申请文件:'+file_str+' 成功!')
        else:
            logger.info('文件保存失败!')
    
    def create_nstd_mat_table(self, nstd_id, res):
        logger.info('正在保存非标物料到维护列表中...')
        self.no_need_mats=[]
        try:
            nstd_app_head.get(nstd_app_head.nstd_app == nstd_id)
            logger.warning('非标申请:'+nstd_id+'已经存在，故未重新创建!')        
            #q= nstd_app_head.update(project=res['project_id'], contract=res['contract'], index_mat=res['index_mat_id'], app_person=res['app_person']).where(nstd_app_head.nstd_app == nstd_id)
            #q.execute()
        except nstd_app_head.DoesNotExist:
            nstd_app_head.create(nstd_app=nstd_id, project=res['project_id'], contract=res['contract'], index_mat=res['index_mat_id'], app_person=res['app_person'])
        
        wbs_list = res['units']

        for wbs in wbs_list:  
            if  len(wbs.strip())==0 and len(wbs_list)>1:
                continue
            nstd_app_link.get_or_create(nstd_app=nstd_id, wbs_no=wbs, mbom_fin=False) 

        for mat in self.nstd_mat_list:
            line = self.mat_items[mat]
            try:
                r=nstd_app_head.select().join(nstd_mat_table).where(nstd_mat_table.mat_no==mat).get()
                nstd_app = none2str(r.nstd_app)
                logger.error('非标物料:'+mat+'已经在非标申请:'+nstd_app+'中提交，请勿重复提交！')
                if nstd_id != nstd_app and mat not in self.no_need_mats:
                    self.no_need_mats.append(mat)
                    self.nstd_mat_list.remove(mat)
            except nstd_mat_table.DoesNotExist:
                nstd_mat_table.create(mat_no=mat, mat_name_cn=line[mat_heads[2]],
                                      mat_name_en=line[mat_heads[3]], drawing_no=line[mat_heads[4]],
                                      mat_unit=line[mat_heads[6]],comments=line[mat_heads[9]],
                                      rp='',box_code_sj='',
                                      nstd_app = nstd_id, mat_app_person=res['app_person'])

            try:
                nstd_mat_fin.get(nstd_mat_fin.mat_no==mat)
            except nstd_mat_fin.DoesNotExist:
                nstd_mat_fin.create(mat_no=mat,justify=-1, mbom_fin=False,\
                                    pu_price_fin=False, co_run_fin=False, modify_by=login_info['uid'], modify_on=datetime.datetime.now())
        
        if len(self.nstd_mat_list)==0:
            logger.error(' 所有非标物料已经在另外的非标申请中提交，请勿重复提交!') 
            return False
        else:      
            logger.info('非标物料维护列表保存进程完成.')
            return True
    
    def fill_nstd_app_table(self, ws, page, nstd, res, count):
        ws.cell(row=6, column=2).value = nstd
        ws.cell(row=7, column=4).value = res['project_name']
        ws.cell(row=7, column=20).value= res['contract']
        wbses = res['units']
        ws.cell(row=7, column=12).value = self.combine_wbs(wbses)
        
        if page==1 and count<=28:
            ran=count
        elif (count%((page-1)*28)>28 and page>1) or (page==1 and count>28):
            ran = 28
        else:
            ran = count%(page*28)
        
        for i in range(1, ran+1):
            mat = self.nstd_mat_list[((page-1)*28+i-1)]
            line = self.mat_items[mat]
            ws.cell(row=i+10, column=3).value = line[mat_heads[2]]
            ws.cell(row=i+10, column=4).value = line[mat_heads[3]]
            ws.cell(row=i+10, column=5).value = mat
            drawing_id = line[mat_heads[4]]
            ws.cell(row=i+10, column=7).value = drawing_id
            ws.cell(row=i+10, column=10).value = line[mat_heads[9]]
            ws.cell(row=i+10, column=20).value = line[mat_heads[6]]
            
            if drawing_id=='NO' or len(drawing_id)==0:
                ws.cell(row=i+10, column=21).value ='否'
            else:
                ws.cell(row=i+10, column=21).value = '是'
                
            if mat in self.mat_tops.keys():
                rp_box = self.mat_tops[mat]['rp_box']
                if not rp_box:
                    ws.cell(row=i+10, column=15).value = rp_box[1]
                    ws.cell(row=i+10, column=17).value = rp_box[0]
                                  
    def style_worksheet(self, ws):        
        thin = Side(border_style="thin", color="000000")
        dash = Side(border_style="dashed", color="000000")
                      
        other_border = Border(top=dash, left=dash, right=dash)
        self.set_border(ws, 'T5:V5', other_border)

        main_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.set_border(ws, 'A6:V40', main_border)
           
        logo = Image(img=os.path.join(cur_dir(),'logo.png'))
        logo.drawing.top = 0
        logo.drawing.left = 30
        logo.drawing.width=110
        logo.drawing.height=71
        head = Image(img=os.path.join(cur_dir(),'head.png'))
        head.drawing.width = 221
        head.drawing.height = 51
                
        ws.add_image(head,'A2')
        ws.add_image(logo,'T1')
        
        ws.print_area ='A1:V40'
                      
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left=0.24
        ws.page_margins.right = 0.24
        ws.page_margins.top = 0.19
        ws.page_margins.bottom=0.63
        ws.page_margins.header = 0
        ws.page_margins.footer= 0
        
        ws.page_setup.scale = 80 
        ws.sheet_properties.pageSetUpPr.fitToPage = True 
             
        ws.oddFooter.left.text='''Songjiang Plant,ThyssenKrupp Elevator ( Shanghai ) Co., Ltd.
No.2, Xunye Road, Sheshan Subarea, Songjiang Industrial Area, Shanghai
Tel.: +86 (21) 37869898   Fax: +86 (21) 57793363
TKEC.SJ-F-03-03'''
        ws.oddFooter.left.font='TKTypeMedium, Regular' 
        ws.oddFooter.left.size =7 
               
        ws.oddFooter.right.text='项目非标物料汇总表V2.01'
        ws.oddFooter.right.font='宋体, Regular' 
        #ws.oddFooter.right.size =8          
            
    def set_border(self, ws, cell_range, border): 
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)               
        rows = ws[cell_range]

        for cell in rows[-1]:
            cell.border = cell.border + bottom 
                   
        for row in rows:
            r = row[-1]
            r.border = r.border+right
            for cell in row:
                cell.border = cell.border+top+left
                                    
    def combine_wbs(self, li):
        li.sort()
        if len(li)>1:
            head = li[0]
        elif li is None:
            return ''
        elif len(li)==0:
            return ''
        else:
            return li[0]
        
        start = int(li[0][11:])
        j=1
        end = ''
        for i in range(1, len(li)):
            if int(li[i][11:]) == start+j:
                j+=1   
            else:
                if j>1:
                    head=head+'~'+end
                elif len(end)>0:
                    head = head+','+end
                
                if j>1:
                    head=head+','+li[i][11:]
                start=int(li[i][11:])

                j=1 
            end = li[i][11:]
            
        if j>1:
            head=head+'~'+end 
        else:
            head = head+','+end
        
        return head
                                
    def get_rel_nstd_info(self, nstd_id):
        try:
            nstd_result = NonstdAppItem.select(NonstdAppItem.link_list, NonstdAppItemInstance.index_mat, NonstdAppItemInstance.res_engineer, NonstdAppItemInstance.create_emp).join(NonstdAppItemInstance, on=(NonstdAppItem.index == NonstdAppItemInstance.index))\
                .where((NonstdAppItemInstance.nstd_mat_app==nstd_id)&(NonstdAppItem.status>=0)&(NonstdAppItemInstance.status>=0)).naive().get()
        except NonstdAppItem.DoesNotExist:
            return None
        
        res ={}
        
        wbs_res = nstd_result.link_list
        index_mat = nstd_result.index_mat

        try:
            emp_res= SEmployee.get(SEmployee.employee==login_info['uid'])
            app_per = emp_res.name
        except SEmployee.DoesNotExist:
            app_per=''         

        i_pos = index_mat.find('-')
        nstd_app_id = index_mat[0:i_pos]
        try:
            nstd_app_result=NonstdAppHeader.get((NonstdAppHeader.nonstd==nstd_app_id)&(NonstdAppHeader.status>=0))
        except NonstdAppHeader.DoesNotExist:
            return None

        project_id = nstd_app_result.project
        contract_id = nstd_app_result.contract
        
        try:
            p_r = ProjectInfo.get(ProjectInfo.project==project_id)
        except ProjectInfo.DoesNotExist:
            return None
        
        project_name = p_r.project_name
        
        if isinstance(wbs_res, str): 
            wbs_list = wbs_res.split(';')
        else:
            wbs_list=['']
            
        wbses =[]
        
        for wbs in wbs_list:
            if  len(wbs.strip())==0 and len(wbs_list)>1:
                continue
            
            w = wbs.strip()
            w = w[0:14]
            wbses.append(w)
            
        res['units']=wbses
        res['contract']=contract_id
        res['project_id']=project_id
        res['project_name']=project_name
        res['app_person'] = app_per
        res['index_mat_id'] = index_mat
        
        return res
                     
    def run_check_in_sap(self):
        if self.sap_thread is not None:
            if self.sap_thread.is_alive():
                messagebox.showinfo('提示','正在后台检查SAP非标物料，请等待完成后再点击!')
                return            
        
        self.sap_thread = refresh_thread(self)
        self.sap_thread.setDaemon(True)
        self.sap_thread.start()
        
    def check_in_sap(self):
        self.nstd_mat_list=[]
        self.hibe_mats=[]
                
        logger.info("正在登陆SAP...")
        config = ConfigParser()
        config.read('sapnwrfc.cfg')
        para_conn  = config._sections['connection']
        para_conn['user'] = base64.b64decode(para_conn['user']).decode()
        para_conn['passwd'] = base64.b64decode(para_conn['passwd']).decode()
        mats = self.mat_items.keys()
        
        try:
            conn = pyrfc.Connection(**para_conn)
            
            imp = []
            for mat in mats:
                line = dict(MATNR=mat, WERKS='2101')
                imp.append(line)
            
            logger.info("正在调用RFC函数...")
            result = conn.call('ZAP_PS_MATERIAL_INFO', IT_CE_MARA=imp, CE_SPRAS='1')
            
            std_mats=[]
            for re in result['OT_CE_MARA']:
                std_mats.append(re['MATNR'])
                
                if re['BKLAS']=='3030' and re['MATNR'] not in self.hibe_mats:
                    self.hibe_mats.append(re['MATNR'])
                
            for mat in mats:
                if mat not in std_mats:
                    logger.info("标记非标物料:"+mat)
                    self.nstd_mat_list.append(mat)
                    self.mark_nstd_mat(mat, True)
                else:
                    self.mark_nstd_mat(mat, False)
                    
            logger.info("非标物料确认完成，共计"+str(len(self.nstd_mat_list))+"个非标物料。")
            
        except pyrfc.CommunicationError:
            logger.error("无法连接服务器")
            return -1
        except pyrfc.LogonError:
            logger.error("无法登陆，帐户密码错误！")
            return -1
        except (pyrfc.ABAPApplicationError, pyrfc.ABAPRuntimeError):
            logger.error("函数执行错误。")
            return -1
        
        conn.close()
                   
        return len(self.nstd_mat_list)
        
    def mark_nstd_mat(self, mat, non=True):
        re=mat_info.get(mat_info.mat_no == mat)
        
        if re.is_nonstd == non:
            return 0
        else:
            q = mat_info.update(is_nonstd=non).where(mat_info.mat_no==mat)
            r = q.execute()
            if(r>0):
                self.change_log('mat_info', 'is_nonstd',mat , (not non), non)
                
            return r

    def search(self,event=None):
        if len(self.find_var.get())==0:
            logger.warning("物料号不能为空，请务必填写物料号")
            return
        
        self.mat_tops = {}        
        self.mat_items={}
        self.mat_list = {}
        self.bom_items = [] 
        self.nstd_mat_list = []
        
        for row in self.mat_tree.get_children():
            self.mat_tree.delete(row) 
        
        logger.info('开始搜索匹配的物料号...')
        if len(self.version_var.get())==0:
            res=mat_info.select(mat_info, bom_header.bom_id, bom_header.revision,bom_header.is_active).join(bom_header, on=(bom_header.mat_no==mat_info.mat_no)).where((mat_info.mat_no.contains(self.find_var.get())) & (bom_header.is_active==True))\
                .order_by(mat_info.mat_no.asc()).naive()  
        else:                 
            res=mat_info.select(mat_info, bom_header.bom_id, bom_header.revision,bom_header.is_active).join(bom_header, on=(bom_header.mat_no==mat_info.mat_no)).where((mat_info.mat_no.contains(self.find_var.get())) & (bom_header.revision==self.version_var.get())& (bom_header.is_active==True))\
                .order_by(mat_info.mat_no.asc()).naive()
            
        if not res:
            logger.warning("没有与搜索条件匹配的物料BOM.")
            return 
        
        for l in res:
            line = {}
            re = {}

            mat = none2str(l.mat_no)
            rev = none2str(l.revision)
                        
            line[mat_heads[0]]= rev
            line[mat_heads[1]]= mat
            line[mat_heads[2]]= none2str(l.mat_name_cn)
            line[mat_heads[3]]= none2str(l.mat_name_en)
            line[mat_heads[4]]= none2str(l.drawing_no)
            line[mat_heads[5]]= 0
            line[mat_heads[6]]= none2str(l.mat_unit)
            line[mat_heads[7]]= none2str(l.mat_material)
            line[mat_heads[8]]= none2str(l.part_weight)
            line[mat_heads[9]]= '' 
            
            revision = none2str(l.rp)
            struct_code = none2str(l.box_code_sj)
                 
            if len(struct_code)>0 and mat not in self.mat_tops:
                re['revision']=revision
                re['struct_code']=struct_code
                rp_box=[]
                if len(none2str(l.rp))!=0 or len(none2str(l.box_code_sj))!=0:
                    rp_box.append(none2str(l.rp))
                    rp_box.append(none2str(l.box_code_sj))
                else:
                    rp_box=None
                re['rp_box'] = rp_box
                self.mat_tops[mat]=re
                
            is_nstd = l.is_nonstd
            if is_nstd and mat not in self.nstd_mat_list:
                self.nstd_mat_list.append(mat)
                    
            if mat not in self.mat_items.keys():
                self.mat_items[mat]=line
            
            item = self.mat_tree.insert('', END, values=dict2list(line))
            
            self.mat_list[item]=line
            if self.get_sub_bom(item, mat, rev):
                self.bom_items.append(item) 
       
        logger.info('正在与SAP匹配确认非标物料，请勿进行其他操作！')
        self.run_check_in_sap()         
                     
    def get_sub_bom(self,item, mat, rev=''):
        r = bom_header.select(bom_header, bom_item, mat_info).join(bom_item, on=(bom_header.bom_id==bom_item.bom_id)).switch(bom_item).join(mat_info, on=(bom_item.component==mat_info.mat_no))\
              .where((bom_header.mat_no==mat)&(bom_header.revision==rev)&(bom_header.is_active==True)).order_by(bom_item.index.asc()).naive()
        
        if not r:
            return False
        
        logger.info('开始搜索物料:'+mat+'的下层BOM')
        for l in r:
            line = {}
            re={}
            
            line[mat_heads[0]]= none2str(l.st_no)
            mat = none2str(l.component)
            line[mat_heads[1]]= mat
            line[mat_heads[2]]= none2str(l.mat_name_cn)
            line[mat_heads[3]]= none2str(l.mat_name_en)
            line[mat_heads[4]]= none2str(l.drawing_no)
            line[mat_heads[5]]= l.qty
            line[mat_heads[6]]= none2str(l.mat_unit)
            line[mat_heads[7]]= none2str(l.mat_material)
            line[mat_heads[8]]= none2str(l.part_weight)
            line[mat_heads[9]]= none2str(l.bom_remark)
            
            is_nstd = l.is_nonstd
            if is_nstd and mat not in self.nstd_mat_list:
                self.nstd_mat_list.append(mat)
                
            tree_item = self.mat_tree.insert(item, END, values=dict2list(line))
            self.mat_list[tree_item]=line
                   
            if mat not in self.mat_items.keys():
                self.mat_items[mat]=line
            
            if self.get_sub_bom(tree_item, mat):
                self.bom_items.append(tree_item)
        
        logger.info('构建物料:'+mat+'下层BOM完成!')        
        return True
                         
    def check_sub_bom(self, mat, ver=''):
        try:
            bom_header.get((bom_header.mat_no==mat)&(bom_header.revision==ver))
            return True
        except bom_header.DoesNotExist:
            return False        
                                         
    def excel_import(self):
        file_list = filedialog.askopenfilenames(title="导入文件", filetypes=[('excel file','.xlsx'),('excel file','.xlsm')])
        if not file_list:
            return

        self.mat_list = {}
        self.mat_pos = 0
        self.mat_tops = {}
        
        self.mat_items={}
        
        for row in self.mat_tree.get_children():
            self.mat_tree.delete(row)
             
        #for node in self.bom_tree.children(0):
            #self.bom_tree.remove_node(node.identifier)

        for file in file_list:
            logger.info("正在读取文件:"+file+",转换保存物料信息,同时构建数据Model")
            c=self.read_excel_files(file)
            logger.info("文件:"+file+"读取完成, 共计处理 "+str(c)+" 个物料。")
            
        #df = pd.DataFrame(self.mat_list,index=mat_heads, columns=[ i for i in range(1, self.mat_pos+1)])
        #model = TableModel(dataframe=df.T)
        #self.mat_table.updateModel(model)
        #self.mat_table.redraw()
        
        logger.info("正在生成BOM层次结构...")
        c = self.build_tree_struct()
        logger.info("Bom结构生成完成，共为"+str(c)+"个发运层物料生成BOM.")
        
        logger.info("正在保存BOM...")
        c = self.save_mats_bom()
        logger.info("共保存"+str(c)+"个物料BOM")
        
        logger.info("正在核查非标物料...")
        self.run_check_in_sap()
                      
    def save_mat_info(self,method=False,**para):
        b_level=False
        
        if para['mat_no'] in self.mat_tops.keys():
            rp_box=  self.mat_tops[para['mat_no']]['rp_box']
            if not rp_box:
                b_level=True
            
        try:
            mat_info.get(mat_info.mat_no == para['mat_no'])
            if method:
                if b_level:
                    q = mat_info.update(mat_name_en=para['mat_name_en'], mat_name_cn=para['mat_name_cn'], drawing_no=para['drawing_no'],mat_material=para['mat_material'],mat_unit=para['mat_unit'],rp=rp_box[0], box_cod_sj=rp_box[1],\
                                mat_material_en=para['mat_material_en'],part_weight=para['part_weight'],comments=para['comments'],modify_by=login_info['uid'],modify_on=datetime.datetime.now()).where(mat_info.mat_no==para['mat_no']) 
                else:
                    q = mat_info.update(mat_name_en=para['mat_name_en'], mat_name_cn=para['mat_name_cn'], drawing_no=para['drawing_no'],mat_material=para['mat_material'],mat_unit=para['mat_unit'],\
                                mat_material_en=para['mat_material_en'],part_weight=para['part_weight'],comments=para['comments'],modify_by=login_info['uid'],modify_on=datetime.datetime.now()).where(mat_info.mat_no==para['mat_no'])
                return q.execute()                          
        except mat_info.DoesNotExist:
            if b_level:
                q = mat_info.insert(mat_no=para['mat_no'], mat_name_en=para['mat_name_en'], mat_name_cn=para['mat_name_cn'], drawing_no=para['drawing_no'],mat_material=para['mat_material'],mat_unit=para['mat_unit'],\
                                mat_material_en=para['mat_material_en'],part_weight=para['part_weight'],rp=rp_box[0],box_code_sj=rp_box[1],comments=para['comments'],modify_by=login_info['uid'],modify_on=datetime.datetime.now())
            else:
                q = mat_info.insert(mat_no=para['mat_no'], mat_name_en=para['mat_name_en'], mat_name_cn=para['mat_name_cn'], drawing_no=para['drawing_no'],mat_material=para['mat_material'],mat_unit=para['mat_unit'],\
                                mat_material_en=para['mat_material_en'],part_weight=para['part_weight'],comments=para['comments'],modify_by=login_info['uid'],modify_on=datetime.datetime.now())
            return q.execute()
        
        return 0
    
    def check_branch(self, item):
        mat = self.mat_tree.item(item, "values")[1]
        for li in self.bom_items:
            if mat == self.mat_tree.item(li, "values")[1]:
                return False
            
        self.bom_items.append(item)
        
        return True
        
    def save_bom_list(self, item):
        it_list = self.mat_tree.item(item, "values")
        mat = it_list[1]
        drawing = it_list[4]
        
        if mat in self.mat_tops.keys():
            revision = self.mat_tops[mat]['revision']
            st_code = self.mat_tops[mat]['struct_code']
        else:
            revision=''
            st_code=''
        
        try:    
            bom_header.get((bom_header.mat_no==mat) & (bom_header.revision==revision) & (bom_header.is_active==True))
            logger.warning(mat+"BOM已经存在，无需重新创建!")
            return 0
        except bom_header.DoesNotExist:
            b_id = self.bom_id_generator()
            q=bom_header.insert(bom_id=b_id, mat_no=mat, revision=revision, drawing_no=drawing, struct_code=st_code,is_active=True,plant=login_info['plant'],\
                                modify_by=login_info['uid'],modify_on=datetime.datetime.now(), create_by=login_info['uid'],create_on=datetime.datetime.now())
            q.execute()
            
        children = self.mat_tree.get_children(item)
        
        data = []
        for child in children:
            d_line = {}
            d_line['bom_id']=b_id
            d_line['index'] = int(self.mat_tree.item(child, "values")[0])
            d_line['st_no'] = self.mat_tree.item(child, "values")[0]
            d_line['component'] = self.mat_tree.item(child,"values")[1]
            d_line['qty']= Decimal(self.mat_tree.item(child,"values")[5])
            d_line['bom_remark']=self.mat_tree.item(child,"values")[9]
            d_line['parent_mat'] = mat
            d_line['modify_by']=login_info['uid']
            d_line['modify_on']= datetime.datetime.now()
            d_line['create_by']=login_info['uid']
            d_line['create_on']=datetime.datetime.now()
            
            data.append(d_line)
            
        q=bom_item.insert_many(data)
        return q.execute()
                        
    def get_rp_boxid(self, struct):
        rp_box = []
        
        try:
            r=struct_gc_rel.get(struct_gc_rel.st_code==struct, struct_gc_rel.plant==login_info['plant'])
        except struct_gc_rel.DoesNotExist:
            return None
        
        rp_box.append(r.rp)
        rp_box.append(r.box_code)
        
        return rp_box
    
    def save_mats_bom(self):
        if len(self.bom_items)==0:
            return 0
        
        i=0
        for item in self.bom_items:
            if self.save_bom_list(item)>0:
                i+=1
        
        return i    
    
    def build_tree_struct(self):
        self.bom_items=[]
        if len(self.mat_list)==0:
            return 0
        
        cur_level = 0
        pre_level = 0
        parent_node = self.mat_tree.insert('', END, values = dict2list(self.mat_list[1]))
        counter =0
        cur_node = parent_node
        self.check_branch(parent_node)
        
        self.mat_tree.item(parent_node, open=True)
        
        for i in range(1,self.mat_pos+1):
            cur_level = tree_level(self.mat_list[i][mat_heads[0]])
            if cur_level==0:
                counter+=1
                
            if (pre_level == cur_level) and pre_level !=0:
                cur_node = self.mat_tree.insert(parent_node, END,  values=dict2list(self.mat_list[i]))
                
            if pre_level<cur_level:
                parent_node = cur_node
                self.check_branch(parent_node)
                cur_node=self.mat_tree.insert(parent_node, END, values=dict2list(self.mat_list[i]))
                
            if pre_level>cur_level:
                while pre_level >= cur_level:
                    parent_node = self.mat_tree.parent(parent_node)
                    if pre_level!=0:
                        pre_level = tree_level(self.mat_tree.item(parent_node, 'values')[0])
                    else:
                        pre_level=-1
                    
                cur_node=self.mat_tree.insert(parent_node, END, values=dict2list(self.mat_list[i]))
                
                if cur_level==0:
                    self.mat_tree.item(cur_node, open=True)
                
            pre_level = cur_level
                
        return counter

    '''        
    def build_tree_struct(self):
        if len(self.mat_list)==0:
            return
        
        cur_level=0
        pre_level=0
        parent_node=0
        counter=0
        for i in range(1, self.mat_pos+1):
            cur_level = tree_level(self.mat_list[i][mat_heads[0]])
            if cur_level==0:
                counter+=1
                
            if pre_level == cur_level:
                self.bom_tree.create_node(i,i,parent_node)
                
            if pre_level < cur_level:
                parent_node = i-1
                self.bom_tree.create_node(i,i,parent_node)
                
            if pre_level > cur_level:
                while pre_level > cur_level:
                    parent_node = self.bom_tree.parent(parent_node).identifier
                    pre_level = tree_level(self.mat_list[parent_node][mat_heads[0]])
                    
                self.bom_tree.create_node(i,i,parent_node)
                                 
            pre_level = cur_level
        
        return counter
    ''' 
                     
    def read_excel_files(self, file):
        '''
                返回值：
                -2: 读取EXCEL失败
                -1 : 头物料位置为空
                0： 头物料的版本已经存在
                1： 
       
        '''
        wb = load_workbook(file, read_only=True,data_only=True)
        sheetnames=wb.get_sheet_names()
        
        if len(sheetnames)==0:
            return -2
        
        counter=0
    
        for i in range(0,len(sheetnames)): 
            if not str(sheetnames[i]).isdigit():
                continue
            
            for j in range(1,19):                         
                mat_line = {}
                mat_top_line={}
                mat=''
                ws = wb.get_sheet_by_name(sheetnames[i]) 
            
                mat_line[mat_heads[0]]=cell2str(ws.cell(row=2*j+1,column=2).value)
                mat = cell2str(ws.cell(row=2*j+1,column=5).value)
                
                if len(mat)==0:
                    break
                
                mat_line[mat_heads[1]]= mat        
                mat_line[mat_heads[2]]=cell2str(ws.cell(row=2*j+1,column=7).value)
                mat_line[mat_heads[3]]=cell2str(ws.cell(row=2*j+2,column=7).value)
                mat_line[mat_heads[4]] = cell2str(ws.cell(row=2*j+1, column=6).value)
                
                qty = cell2str(ws.cell(row=2*j+1,column=3).value)
                if len(qty)==0: 
                    continue
                
                self.mat_pos+=1
                counter+=1
                
                mat_line[mat_heads[5]] = Decimal(qty)
                mat_line[mat_heads[6]]=cell2str(ws.cell(row=2*j+1,column=4).value)
                mat_line[mat_heads[7]] = cell2str(ws.cell(row=2*j+1,column=9).value)
                material_en = cell2str(ws.cell(row=2*j+2, column=9).value)
                
                weight = cell2str(ws.cell(row=2*j+1, column=10).value)
                if len(weight)==0:
                    mat_line[mat_heads[8]]=0
                else:
                    mat_line[mat_heads[8]]=Decimal(weight)
                    
                mat_line[mat_heads[9]]=cell2str(ws.cell(row=2*j+1, column=11).value)
                
                if sheetnames[i]=='1' and j==1:
                    mat_top_line['revision'] = cell2str(ws.cell(row=43, column=8).value)
                    mat_top_line['struct_code']=cell2str(ws.cell(row=39,column=12).value)
                    rp_box = self.get_rp_boxid(mat_top_line['struct_code'])
                    mat_top_line['rp_box'] = rp_box
                    
                    self.mat_tops[mat_line[mat_heads[1]]]=mat_top_line
                
                #保存物料基本信息
                if self.save_mat_info(mat_no=mat_line[mat_heads[1]], mat_name_en=mat_line[mat_heads[3]], mat_name_cn=mat_line[mat_heads[2]], drawing_no=mat_line[mat_heads[4]],mat_material=mat_line[mat_heads[7]],mat_unit=mat_line[mat_heads[6]],\
                                mat_material_en=material_en,part_weight=mat_line[mat_heads[8]],comments=mat_line[mat_heads[9]])==0:
                    logger.info(mat_line[mat_heads[1]]+'数据库中已经存在,故没有保存')
                else:
                    logger.info(mat_line[mat_heads[1]]+'保存成功。')
                
                self.mat_list[self.mat_pos] = mat_line
                
                if mat not in self.mat_items.keys():
                    self.mat_items[mat] = mat_line
                               
        return counter
                                                    
    def bom_id_generator(self):
        try:
            bom_res = id_generator.get(id_generator.id == 1)
        except id_generator.DoesNotExist:
            return None
        
        pre_char = none2str(bom_res.pre_character)
        fol_char = none2str(bom_res.fol_character)
        c_len = bom_res.id_length
        cur_id = bom_res.current
        step = bom_res.step        
        new_id=str(cur_id+step)
        #前缀+前侧补零后长度为c_len+后缀, 组成新的BOM id               
        id_char = pre_char+new_id.zfill(c_len)+fol_char
        
        q=id_generator.update(current=cur_id+step).where(id_generator.id==1)
        q.execute()
        
        return id_char
    
    def change_log(self,table,section,key, old,new):
        q = s_change_log.insert(table_name=table,change_section=section,key_word=str(key),old_value=str(old),new_value=str(new),log_on=datetime.datetime.now(), log_by=login_info['uid'] )
        q.execute()
                              
class Application():
    def __init__(self, root):      
        main_frame = mainframe(root)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.grid(row=0, column=0, sticky=NSEW)
        LoginForm(root,'用户登陆')
        main_frame.display_widgets()
        root.protocol("WM_DELETE_WINDOW", self.quit_func)

    def quit_func(self):          
        if pg_db.get_conn():
            pg_db.close()
            
        if mbom_db.get_conn() and login_info['status']:
            log_loger = login_log.update(logout_time = datetime.datetime.now(), log_status=False).where((login_log.employee==login_info['uid'])&(login_log.log_status==True))
            log_loger.execute()            
            mbom_db.close()
        root.destroy()
                
if __name__ == '__main__':   
    root=Tk() 
    #root.resizable(0, 0)
    
    #root.attributes("-zoomed", 1) # this line removes the window managers bar       
        
    try:                                   # Automatic zoom if possible
        root.wm_state("zoomed")
    except TclError:                    # Manual zoom
        # get the screen dimensions
        width = root.winfo_screenwidth()
        height = root.winfo_screenheight()

        # borm: width x height + x_offset + y_offset
        geom_string = "%dx%d+0+0" % (width, height)
        root.wm_geometry(geom_string)
        
    root.title(NAME+PUBLISH_KEY+VERSION)
    default_font = font.nametofont("TkDefaultFont")
    default_font.configure(size=10)  
    root.option_add("*Font", default_font)
    Application(root)
    root.geometry('800x600')
    root.mainloop()