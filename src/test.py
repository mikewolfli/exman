#coding=utf-8
'''
Created on 2016年8月26日

@author: 10256603
'''
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

wb = Workbook()
ws = wb.active
my_cell = ws['B2']
my_cell.value = "My Cell"
thin = Side(border_style="thin", color="000000")
dot = Side(border_style="dashDot", color="000000")
thick = Side(border_style="thick", color="000000")
dash = Side(border_style="dashed", color="000000")

border = Border(top=dot, left=thin, right=thick, bottom=dash)
fill = PatternFill("solid", fgColor="DDDDDD")
fill = GradientFill(stop=("000000", "FFFFFF"))
font = Font(b=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")


style_range(ws, 'B2:F4', border=border, fill=fill, font=font, alignment=al)

ws.page_margins.left= float(u'0.24')
ws.page_margins.right = float(u'0.24')
ws.page_margins.top = float(u'0.19')
ws.page_margins.bottom=float(u'0.63')
ws.page_margins.header = float(u'0.0')
ws.page_margins.footer= float(u'0.0')

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4

ws.oddFooter.left.font='TKTypeMedium, Regular' 
ws.oddFooter.left.size =7         
ws.oddFooter.left.text ='''Songjiang Plant,ThyssenKrupp Elevator ( Shanghai ) Co., Ltd.
No.2, Xunye Road, Sheshan Subarea, Songjiang Industrial Area, Shanghai
Tel.: +86 (21) 37869898   Fax: +86 (21) 57793363
TKEC.SJ-F-03-03'''

ws.oddFooter.right.text='项目非标物料汇总表V2.01'
ws.oddFooter.right.font='宋体, Regular' 
#ws.oddFooter.right.size =8  

wb.save("m:/styled9.xlsx")

